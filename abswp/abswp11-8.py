import openpyxl
import os
#lesson 2 days ago:

os.chdir('c:\\Users\\rimuru\\Documents')

workbook = openpyxl.load_workbook('example.xlsx')
print(type(workbook))
sheet = workbook['Sheet']
print(type(sheet))
print(workbook.sheetnames)

cell = sheet['A1']
print(cell.value)
print(str(cell.value))
print(str(sheet['A1']))
#print(sheet.cell(row=1,column=2)
print(sheet['B1'])
for i in range(1,8):
    print(i,sheet.cell(row=i,column=2))

#-------------------------------------------------------
os.chdir('c:\\Users\\rimuru\\Documents')
wb = openpyxl.Workbook()
print(wb)

wb.get_sheet_names()
sheet = wb.get_sheet_by_name('Sheet')
print(sheet)
sheet["A1"].value
print(sheet["A1"].value == None)

sheet['A1'] = 42
sheet['A2'] = 'Hello'


sheet2 = wb.create_sheet()
print(wb.get_sheet_names())
print(sheet2.title)
sheet2.title = 'My New Sheet Name'
print(wb.get_sheet_names())
#wb.save('example2.xlsx')
wb.create_sheet(index = 0, title = 'My Other Sheet')
#wb.save('example3.xlsx')

